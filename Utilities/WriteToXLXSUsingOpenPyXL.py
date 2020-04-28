import openpyxl

# Create an object of workbook
objWK = openpyxl.Workbook()

# Get the active sheet and name it
objWS = objWK.active
objWS.title = "Name"
print(objWS.title)
# Add values to the cell
c1 = objWS.cell(1, 1)
c1.value = "First name"
c2 = objWS.cell(1, 2)
c2.value = "Sadashiva"
c3 = objWS.cell(2, 1)
c3.value = "Last name"
c4 = objWS.cell(2, 2)
c4.value = "Ashok"


# Create a new sheet
objWK.create_sheet("Contact")
objWS = objWK["Contact"]
c1 = objWS.cell(1, 1)
c1.value = "Email:"
c2 = objWS.cell(1, 2)
c2.value = "sadacs2005@gmail.com"


objWK.save(".././WriteExcel2.xlsx")