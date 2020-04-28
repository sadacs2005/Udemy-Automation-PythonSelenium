import openpyxl

try:
    # Create an object of Workbook
    objWorkBook = openpyxl.load_workbook(".././ReadExcel.xlsx")

    # Create an object of the worksheet
    objWorkSheet = objWorkBook["Sheet1"]

    # Get number of rows
    rows = objWorkSheet.max_row
    # Get number of columns
    columns = objWorkSheet.max_column

    for i in range(2, rows + 1):
        for j in range(1, columns + 1):
            cell = objWorkSheet.cell(i, j)
            print(cell.value)

except FileNotFoundError as e:
    print("File not found: " + str(e))


def readFromExcel():
    try:
        wb = openpyxl.load_workbook(".././ReadExcel.xlsx")
        ws = wb["Sheet1"]
        num_of_rows = ws.max_row
        innerList = []
        mainList = []
        for i in range(2, num_of_rows + 1):
            innerList = []
            uname = ws.cell(i, 1)
            pname = ws.cell(i, 2)
            innerList.insert(0, uname.value)
            innerList.insert(1, pname.value)
            mainList.insert(i - 1, innerList)
        return mainList
    except FileNotFoundError as e:
        print("File not found: " + str(e))
